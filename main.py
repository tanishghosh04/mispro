import streamlit as st
from bs4 import BeautifulSoup
from openpyxl import Workbook
import json
import datetime
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
import requests
from io import BytesIO


def check_saturday(date_str):
    # date_str format: "YYYY-MM-DD"
    date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()

    # Check if it's a Saturday
    if date_obj.weekday() != 5:
        return "Not a Saturday"

    # Count Saturdays in month up to this date
    count = 0
    for day in range(1, date_obj.day + 1):
        d = datetime.date(date_obj.year, date_obj.month, day)
        if d.weekday() == 5:  # Saturday
            count += 1

    if count == 2:
        return "This is the 2nd Saturday"
    elif count == 4:
        return "This is the 4th Saturday"
    else:
        return f"This is the {count}th Saturday (not 2nd or 4th)"


def download_excel():
    session = requests.Session()

    # Step 1: Load login page (important for session + CSRF cookies)
    login_page_url = "https://www.thawksolution.com/home/login"
    session.get(login_page_url)

    # Step 2: Login request
    login_api_url = "https://www.thawksolution.com/home/LoginResponse"
    payload = {
        "email": emailID,
        "password": password
    }

    headers = {
        "User-Agent": "Mozilla/5.0",
        "X-Requested-With": "XMLHttpRequest"
    }

    resp = session.post(login_api_url, data=payload, headers=headers)
    print("Login status:", resp.status_code)

    
    print("\nCookies after login:")
    for c in session.cookies:
        print(c.name, "=", c.value)

    # Step 3: Call the internal API with your date
    internal_api_url = "https://www.thawksolution.com/Employee/TodayAttendancePartial2"
    data_payload = {
        "date": fromDate
    }

    response = session.post(internal_api_url, data=data_payload, headers=headers)

    print("\nInternal API Status:", response.status_code)
    content = response.text

    if response.status_code == 500:
        st.error("Wrong Credentials!")
        return None
    else:
        st.success("Login successful")

        # print(jcontent)
        soup = BeautifulSoup(content, 'html.parser')
        # a = soup.find_all('tr')[4].find_all('td')
        a = soup.find_all('tr')
        # print(a)

        # Creating instance of MIS Workbook
        wb = Workbook()
        ws = wb.active

        # Merging Cells on Holidays
        def mergeCellsForHolidays():
            ws[f"D{i + pointerValue}"] = "Holiday"
            ws.merge_cells(f"D{i + pointerValue}:J{i + pointerValue}")
            ws[f"D{i + pointerValue}"].alignment = Alignment(horizontal='center')

        ws['B2'] = "Name"
        ws['C2'] = a[3].find_all('td')[1].text
        ws['B3'] = "HOD"
        ws['C3'] = hodName

        # Merging the cells
        ws.merge_cells('B5:J5')
        ws[
            'B5'] = f"MIS for the Month of {datetime.datetime(int(fromDate[0:4]), int(fromDate[5:7]), int(fromDate[8:10])).strftime('%B')}"
        ws['B5'].alignment = Alignment(horizontal='center')
        ws['B5'].font = Font(bold=True)

        ws['B6'] = "Date"
        ws['C6'] = "Day"
        ws['D6'] = "Time IN"
        ws['E6'] = "Time OUT"
        ws['F6'] = "Client"
        ws['G6'] = "Personal/Client's /SKA's Laptop"
        ws['H6'] = "WORK IN DETAIL"
        ws['I6'] = "Credit Days (Worked on Holidays)"
        ws['J6'] = "Remarks"

        # Laptop Detail
        laptop = laptopInput

        # Count of Saturday, Credit Days
        saturdayCount = 0
        creditDays = 0

        # Adding Border
        # border_thin = Side(style='thin')

        total_rows = len(a)
        if total_rows <= 2:
            st.warning("No attendance data found!")
            return None



        for i in range(len(a)):

            if i <= 1:
                pass
            else:
                pointerValue = 5
                row = a[i].find_all('td')
                date = str(row[0].text)
                timeIn = row[3].text.strip()
                timeOut = row[7].text

                # Stops MIS report at (eDate +1)
                if date == f"{toDate[8:10]}/{toDate[5:7]}/{toDate[0:4]}":
                    break

                # Writes value to desired cells
                ws['C' + str(i + pointerValue)] = datetime.date(int(date[6:10]), int(date[3:5]),
                                                                int(date[0:2])).strftime("%A")
                ws['D' + str(i + pointerValue)] = timeIn
                ws['E' + str(i + pointerValue)] = timeOut
                ws['G' + str(i + pointerValue)] = laptop

                for hdate in jcontent['holiday-list']:
                    if hdate == date:
                        mergeCellsForHolidays()
                    else:
                        pass

                ws['B' + str(i + pointerValue)] = date

                # Declaring Sunday as Holiday and instance for credit days
                if datetime.date(int(date[6:10]), int(date[3:5]), int(date[0:2])).strftime("%A") == "Sunday":
                    if timeIn == "":
                        mergeCellsForHolidays()
                    else:
                        creditDays += 1
                        ws['I' + str(i + pointerValue)] = 1
                        ws['I' + str(i + pointerValue)].alignment = Alignment(horizontal='center')

                # Declaring Second & Fourth Saturday as Holiday and instance for credit days
                dayCheck = check_saturday(f"{date[6:10]}-{date[3:5]}-{date[0:2]}")
                if dayCheck == "This is the 2nd Saturday" or dayCheck == "This is the 4th Saturday":
                    if timeIn == "":
                        mergeCellsForHolidays()
                    else:
                        creditDays += 1
                        ws['I' + str(i + pointerValue)] = 1
                        ws['I' + str(i + pointerValue)].alignment = Alignment(horizontal='center')


        # Declaring Fill Color
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Applies border to all cells
        rangeBorder = ws['B6':'J' + str(ws.max_row)]
        for cell in rangeBorder:
            for x in cell:
                x.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))

        for cell in ws['B5':'J5']:
            for x in cell:
                x.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin'))
                x.fill = fill

        for bCell in ws['B6':'J6']:
            for cell in bCell:
                cell.font = Font(bold=True)
                cell.fill = fill

        for bTCell in ws['B2':'C3']:  # Making NAME and HOD Bold
            for cell in bTCell:
                cell.font = Font(bold=True)

        # Autofit Column Width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[column].width = max_length + 2

    # SAVE EXCEL IN MEMORY (NOT ON DISK)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# For presenting in browser
st.title("Online MIS Generator")
emailID = st.text_input("Enter your Email ID", key="emailID")
password = st.text_input("Enter your Password", key="password", type="password")
hodName = st.text_input("Name of HOD", key="hodname")
laptopInput = st.selectbox(
    "Laptop (SKA/ Personal)",
    ["SKA", "Personal"]
)
fromDate = str(st.date_input("From Date", key="fromDate"))
toDate = str(st.date_input("To Date (+ 1)", key="toDate"))
fomattedToDate = toDate[8:10] + '-' + toDate[5:7] + '-' + toDate[0:4]

# Holiday List JSON
with open('holiday.json', 'r') as jfile:
    jcontent = json.load(jfile)

# ✅ Enable download button only when all fields are filled
if emailID and password and hodName and toDate and laptopInput != "None":

    if st.button("Start MIS Generation"):
        buffer = download_excel()

        if buffer:
            st.download_button(
                label="📥 Click here to download MIS Excel File",
                data=buffer,
                file_name=f"MIS_{fromDate}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

else:
    st.warning("⚠️ Please fill all fields before downloading the Excel file.")
    st.button("Download Excel file", disabled=True)







